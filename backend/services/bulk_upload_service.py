import aiohttp
import asyncio
from openpyxl import load_workbook
import logging
from typing import List, Dict
from datetime import datetime, timezone
import tempfile
import os
from .pdf_classifier import PDFClassifier
from .sharepoint_service import SharePointService

logger = logging.getLogger(__name__)

class BulkUploadService:
    def __init__(self, db):
        self.db = db
        self.pdf_classifier = PDFClassifier()
        self.sharepoint_service = SharePointService()
    
    async def process_excel_file(self, job_id: str, file_path: str, manufacturer_name: str, sharepoint_folder: str):
        """Process an Excel file with part numbers and PDF URLs"""
        try:
            logger.info(f"Starting bulk upload for job {job_id}")
            
            # Update job status
            await self._update_job_status(job_id, "processing")
            
            # Check if job was cancelled
            if await self._is_job_cancelled(job_id):
                logger.info(f"Bulk upload job {job_id} was cancelled before processing")
                return
            
            # Parse Excel file
            pdf_items = await self._parse_excel_file(file_path)
            
            logger.info(f"Found {len(pdf_items)} items in Excel file for job {job_id}")
            
            # Update job with found PDFs count
            await self.db.bulk_upload_jobs.update_one(
                {"id": job_id},
                {"$set": {"total_items": len(pdf_items), "updated_at": datetime.now(timezone.utc).isoformat()}}
            )
            
            # Check if job was cancelled
            if await self._is_job_cancelled(job_id):
                logger.info(f"Bulk upload job {job_id} was cancelled after parsing")
                return
            
            # Download and classify PDFs
            await self._update_job_status(job_id, "downloading")
            await self._download_and_classify_pdfs(job_id, pdf_items, manufacturer_name)
            
            # Check if job was cancelled
            if await self._is_job_cancelled(job_id):
                logger.info(f"Bulk upload job {job_id} was cancelled after downloading")
                return
            
            # Upload to SharePoint
            await self._update_job_status(job_id, "uploading")
            await self._upload_to_sharepoint(job_id, sharepoint_folder)
            
            # Mark as completed
            await self._update_job_status(job_id, "completed")
            logger.info(f"Bulk upload job {job_id} completed successfully")
            
        except Exception as e:
            logger.error(f"Error in bulk upload job {job_id}: {str(e)}")
            await self.db.bulk_upload_jobs.update_one(
                {"id": job_id},
                {"$set": {
                    "status": "failed",
                    "error_message": str(e),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
        finally:
            # Clean up temp file
            if os.path.exists(file_path):
                os.unlink(file_path)
    
    async def _parse_excel_file(self, file_path: str) -> List[Dict]:
        """Parse Excel file and extract part numbers and URLs"""
        pdf_items = []
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet = workbook.active
            
            # Skip header row, start from row 2
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2:
                    continue
                
                part_number = row[0]
                pdf_url = row[1]
                
                # Skip empty rows
                if not part_number or not pdf_url:
                    continue
                
                # Validate URL
                if not isinstance(pdf_url, str) or not pdf_url.startswith(('http://', 'https://')):
                    logger.warning(f"Invalid URL for part {part_number}: {pdf_url}")
                    continue
                
                pdf_items.append({
                    "part_number": str(part_number).strip(),
                    "url": str(pdf_url).strip()
                })
            
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            raise Exception(f"Failed to parse Excel file: {str(e)}")
        
        return pdf_items
    
    async def _download_and_classify_pdfs(self, job_id: str, pdf_items: List[Dict], manufacturer_name: str):
        """Download PDFs (no AI classification needed for bulk upload)"""
        processed_count = 0
        failed_count = 0
        
        for item in pdf_items:
            try:
                part_number = item['part_number']
                pdf_url = item['url']
                
                # Download PDF
                async with aiohttp.ClientSession() as session:
                    async with session.get(pdf_url, timeout=aiohttp.ClientTimeout(total=60)) as response:
                        if response.status == 200:
                            # Get filename from URL or use part number
                            filename = pdf_url.split('/')[-1]
                            if not filename.lower().endswith('.pdf'):
                                filename = f"{part_number}.pdf"
                            
                            file_size = int(response.headers.get('content-length', 0))
                            
                            # For bulk upload, all PDFs are assumed to be technical product data
                            # No AI classification needed since user explicitly provided technical data URLs
                            
                            # Save PDF record
                            pdf_record = {
                                "id": str(datetime.now(timezone.utc).timestamp()).replace('.', ''),
                                "job_id": job_id,
                                "part_number": part_number,
                                "filename": filename,
                                "source_url": pdf_url,
                                "file_size": file_size,
                                "is_technical": True,  # Always true for bulk upload
                                "classification_reason": "Bulk upload - user-provided technical product data",
                                "document_type": "Technical Product Data Sheet",
                                "sharepoint_uploaded": False,
                                "sharepoint_id": None,
                                "download_status": "success",
                                "error_message": None,
                                "created_at": datetime.now(timezone.utc).isoformat()
                            }
                            
                            await self.db.bulk_upload_pdfs.insert_one(pdf_record)
                            processed_count += 1
                            
                            logger.info(f"Downloaded PDF for part {part_number}: {filename}")
                        else:
                            # Save failed record
                            failed_count += 1
                            filename = pdf_url.split('/')[-1] if '/' in pdf_url else f"{part_number}.pdf"
                            pdf_record = {
                                "id": str(datetime.now(timezone.utc).timestamp()).replace('.', ''),
                                "job_id": job_id,
                                "part_number": part_number,
                                "filename": filename,
                                "source_url": pdf_url,
                                "file_size": 0,
                                "is_technical": True,
                                "classification_reason": "Bulk upload - user-provided technical product data",
                                "document_type": "Technical Product Data Sheet",
                                "sharepoint_uploaded": False,
                                "sharepoint_id": None,
                                "download_status": "failed",
                                "error_message": f"HTTP {response.status} - URL not accessible",
                                "created_at": datetime.now(timezone.utc).isoformat()
                            }
                            await self.db.bulk_upload_pdfs.insert_one(pdf_record)
                            logger.warning(f"Failed to download PDF for part {part_number}: HTTP {response.status} - {pdf_url}")
            
            except Exception as e:
                failed_count += 1
                logger.error(f"Error processing part {item.get('part_number')}: {str(e)}")
                # Save error record
                pdf_record = {
                    "id": str(datetime.now(timezone.utc).timestamp()).replace('.', ''),
                    "job_id": job_id,
                    "part_number": item.get('part_number', 'Unknown'),
                    "filename": f"{item.get('part_number', 'unknown')}.pdf",
                    "source_url": item.get('url', ''),
                    "file_size": 0,
                    "is_technical": True,
                    "classification_reason": "Bulk upload - user-provided technical product data",
                    "document_type": "Technical Product Data Sheet",
                    "sharepoint_uploaded": False,
                    "sharepoint_id": None,
                    "download_status": "failed",
                    "error_message": str(e),
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await self.db.bulk_upload_pdfs.insert_one(pdf_record)
        
        # Update job with processed and failed counts
        await self.db.bulk_upload_jobs.update_one(
            {"id": job_id},
            {"$set": {
                "total_classified": processed_count,
                "total_failed": failed_count,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    async def _upload_to_sharepoint(self, job_id: str, sharepoint_folder: str):
        """Upload technical PDFs to SharePoint"""
        # Get all technical PDFs for this job
        technical_pdfs = await self.db.bulk_upload_pdfs.find({
            "job_id": job_id,
            "is_technical": True,
            "sharepoint_uploaded": False
        }).to_list(1000)
        
        uploaded_count = 0
        
        for pdf in technical_pdfs:
            try:
                # Download PDF
                async with aiohttp.ClientSession() as session:
                    async with session.get(pdf['source_url'], timeout=aiohttp.ClientTimeout(total=60)) as response:
                        if response.status == 200:
                            pdf_content = await response.read()
                            
                            # Upload to SharePoint
                            sharepoint_id = await self.sharepoint_service.upload_pdf(
                                filename=pdf['filename'],
                                content=pdf_content,
                                folder_path=sharepoint_folder
                            )
                            
                            # Update PDF record
                            await self.db.bulk_upload_pdfs.update_one(
                                {"id": pdf['id']},
                                {"$set": {
                                    "sharepoint_uploaded": True,
                                    "sharepoint_id": sharepoint_id
                                }}
                            )
                            
                            uploaded_count += 1
                            logger.info(f"Uploaded PDF to SharePoint: {pdf['filename']}")
            
            except Exception as e:
                logger.error(f"Error uploading PDF {pdf['filename']} to SharePoint: {str(e)}")
        
        # Update job with uploaded count
        await self.db.bulk_upload_jobs.update_one(
            {"id": job_id},
            {"$set": {"total_uploaded": uploaded_count, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    async def _update_job_status(self, job_id: str, status: str):
        """Update job status"""
        await self.db.bulk_upload_jobs.update_one(
            {"id": job_id},
            {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    async def _is_job_cancelled(self, job_id: str) -> bool:
        """Check if job has been cancelled"""
        job = await self.db.bulk_upload_jobs.find_one({"id": job_id}, {"_id": 0, "status": 1})
        return job and job.get("status") == "cancelled"